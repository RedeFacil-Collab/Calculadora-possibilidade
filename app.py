import io
import json
import os
import secrets
from datetime import timedelta
from functools import wraps
from pathlib import Path

from flask import Flask, Response, g, jsonify, redirect, render_template, request, send_from_directory, send_file, session, stream_with_context, url_for
from googleapiclient.errors import HttpError
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.utils import secure_filename

from auth import AuthStore
from parsing import normalize_bank_factors, parse_table_text, slugify_bank_name
from presence import PresenceStore
from sheets_service import CommercialMatrixService, SheetsConfigurationError

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")
DATA_DIR = BASE_DIR / "data"
DATABASE_URL = os.getenv("DATABASE_URL", "")
REDIS_URL = os.getenv("REDIS_URL", "")
DISCOUNTS_FILE = DATA_DIR / "discounts.json"
BANK_FACTORS_FILE = DATA_DIR / "bank_factors.json"
BANK_LOGOS_FILE = DATA_DIR / "bank_logos.json"
BANK_CATALOG_FILE = DATA_DIR / "bank_catalog.json"
BLOCKED_ENTITIES_FILE = DATA_DIR / "blocked_entities.json"
BANK_LOGOS_DIR = BASE_DIR / "static" / "bank-logos"
ALLOWED_LOGO_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}

DEFAULT_DISCOUNTS = [
    {"name": "BanriSul - Crescente", "percent": 0.0},
    {"name": "BanriSul - Decrescente", "percent": 18.0},
    {"name": "BanriSul - Quitação", "percent": 20.0},
    {"name": "BB - Crescente", "percent": 4.0},
    {"name": "BB - Decrescente", "percent": 10.0},
    {"name": "BB - Quitação", "percent": 12.0},
    {"name": "BMG - Crescente", "percent": 0.0},
    {"name": "BMG - Decrescente", "percent": 12.0},
    {"name": "BMG - Quitação", "percent": 20.0},
    {"name": "Bradesco - Crescente", "percent": 0.0},
    {"name": "Bradesco - Decrescente", "percent": 10.0},
    {"name": "Bradesco - Quitação", "percent": 12.0},
    {"name": "CEF - Crescente", "percent": 0.0},
    {"name": "CEF - Decrescente", "percent": 10.0},
    {"name": "CEF - Quitação", "percent": 12.0},
    {"name": "Daycoval - Crescente", "percent": 0.0},
    {"name": "Daycoval - Decrescente", "percent": 15.0},
    {"name": "Daycoval - Quitação", "percent": 20.0},
    {"name": "Digio - Crescente", "percent": 3.0},
    {"name": "Digio - Decrescente", "percent": 6.0},
    {"name": "Digio - Quitação", "percent": 15.0},
    {"name": "Inter - Crescente", "percent": 0.0},
    {"name": "Inter - Decrescente", "percent": 8.0},
    {"name": "Inter - Quitação", "percent": 12.0},
    {"name": "PAN - Crescente", "percent": 5.0},
    {"name": "PAN - Decrescente", "percent": 12.0},
    {"name": "PAN - Quitação", "percent": 18.0},
    {"name": "Safra - Crescente", "percent": 5.0},
    {"name": "Safra - Decrescente", "percent": 15.0},
    {"name": "Safra - Quitação", "percent": 20.0},
    {"name": "Santander - Crescente", "percent": 3.0},
    {"name": "Santander - Decrescente", "percent": 8.0},
    {"name": "Santander - Quitação", "percent": 15.0},
    {"name": "Sicoob - Crescente", "percent": 0.0},
    {"name": "Sicoob - Decrescente", "percent": 18.0},
    {"name": "Sicoob - Quitação", "percent": 25.0},
    {"name": "Associação - Quitação", "percent": 12.0},
]

DEFAULT_BANK_FACTORS = [
    {"bank": "Safra", "installments": 120, "factor": 0.01893, "active": True, "product": "normal"},
    {"bank": "Safra", "installments": 96, "factor": 0.02253, "active": True, "product": "normal"},
    {"bank": "Pan", "installments": 120, "factor": 0.024001, "active": True, "product": "normal"},
    {"bank": "Pan", "installments": 96, "factor": 0.025335, "active": True, "product": "normal"},
    {"bank": "Pan", "installments": 84, "factor": 0.0288085, "active": True, "product": "normal"},
    {"bank": "Daycoval", "installments": 120, "factor": 0.021692, "active": True, "product": "normal"},
    {"bank": "Daycoval", "installments": 96, "factor": 0.02407, "active": True, "product": "normal"},
    {"bank": "Daycoval", "installments": 84, "factor": 0.025301, "active": True, "product": "normal"},
    {"bank": "Daycoval", "installments": 72, "factor": 0.027843, "active": True, "product": "normal"},
    {"bank": "Daycoval", "installments": 60, "factor": 0.034357, "active": True, "product": "normal"},
    {"bank": "Daycoval", "installments": 48, "factor": 0.038199, "active": True, "product": "normal"},
    {"bank": "Santander", "installments": 120, "factor": 0.023252, "active": True, "product": "normal"},
    {"bank": "Santander", "installments": 96, "factor": 0.02496, "active": True, "product": "normal"},
    {"bank": "Santander", "installments": 84, "factor": 0.026301, "active": True, "product": "normal"},
    {"bank": "Santander", "installments": 72, "factor": 0.028202, "active": True, "product": "normal"},
    {"bank": "Santander", "installments": 60, "factor": 0.03092, "active": True, "product": "normal"},
    {"bank": "Santander", "installments": 48, "factor": 0.035865, "active": True, "product": "normal"},
    {"bank": "Santander", "installments": 36, "factor": 0.04269, "active": True, "product": "normal"},
    {"bank": "Digio", "installments": 120, "factor": 0.020741, "active": True, "product": "normal"},
    {"bank": "Digio", "installments": 96, "factor": 0.023225, "active": True, "product": "normal"},
    {"bank": "Digio", "installments": 84, "factor": 0.027818, "active": True, "product": "normal"},
    {"bank": "Digio", "installments": 72, "factor": 0.02947, "active": True, "product": "normal"},
    {"bank": "Digio", "installments": 60, "factor": 0.03194, "active": True, "product": "normal"},
    {"bank": "Digio", "installments": 48, "factor": 0.035865, "active": True, "product": "normal"},
    {"bank": "Digio", "installments": 36, "factor": 0.04269, "active": True, "product": "normal"},
]

app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.getenv("APP_SECRET_KEY") or os.urandom(32),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.getenv("SESSION_COOKIE_SECURE", "false").lower() == "true",
    PERMANENT_SESSION_LIFETIME=timedelta(days=14),
    MAX_CONTENT_LENGTH=5 * 1024 * 1024,
)
auth_store = AuthStore(DATABASE_URL)
auth_store.initialize()
presence_store = PresenceStore(REDIS_URL)
presence_store.ping()
commercial_matrix_service = CommercialMatrixService()


def request_ip() -> str:
    raw = request.headers.get("X-Forwarded-For") or request.remote_addr or "unknown"
    return raw.split(",")[0].strip()


def api_auth_error(message: str, status: int):
    return jsonify({"error": message}), status


def generate_csrf_token() -> str:
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_urlsafe(32)
    return session["csrf_token"]


def validate_csrf_token() -> str | None:
    token = request.headers.get("X-CSRF-Token") or request.form.get("csrf_token")
    expected = session.get("csrf_token")
    if not expected or not token or not secrets.compare_digest(token, expected):
        return "Token CSRF inválido ou ausente."
    return None


@app.before_request
def enforce_csrf():
    if request.method in ("GET", "HEAD", "OPTIONS"):
        return None
    exempt = ("/api/auth/login",)
    if request.path in exempt:
        return None
    if not session.get("user_id") and not session.get("pending_password_change_user_id"):
        return None
    error = validate_csrf_token()
    if error:
        return jsonify({"error": error}), 403
    return None


@app.before_request
def require_authenticated_user():
    """Aplica o controle de acesso antes das rotas da calculadora."""
    allowed_prefixes = ("/login", "/api/auth/", "/static/")
    if request.path.startswith(allowed_prefixes):
        return None

    user_id = session.get("user_id")
    user = auth_store.get_user(user_id) if user_id else None
    if not user or not user.get("is_active"):
        session.clear()
        if request.path.startswith("/api/"):
            return api_auth_error("Autenticação necessária.", 401)
        return redirect(url_for("login"))

    g.current_user = user
    auth_store.touch_user(user["id"])
    presence_store.mark_online(user["id"])
    if request.path.startswith(("/discounts-editor", "/admin", "/api/admin/")) and user["role"] != "admin":
        return "Acesso restrito a administradores.", 403
    return None


def require_role(*roles: str):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user = getattr(g, "current_user", None)
            if not user or user["role"] not in roles:
                return api_auth_error("Seu perfil não possui essa permissão.", 403)
            return view(*args, **kwargs)
        return wrapped
    return decorator


def asset_url(filename: str) -> str:
    file_path = BASE_DIR / "static" / filename
    version = int(file_path.stat().st_mtime) if file_path.exists() else 0
    return url_for("static", filename=filename, v=version)


@app.after_request
def set_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response


@app.context_processor
def inject_globals():
    return {"asset_url": asset_url, "csrf_token": generate_csrf_token}


def ensure_data_files() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    BANK_LOGOS_DIR.mkdir(parents=True, exist_ok=True)
    if not DISCOUNTS_FILE.exists():
        DISCOUNTS_FILE.write_text(
            json.dumps(DEFAULT_DISCOUNTS, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    if not BANK_FACTORS_FILE.exists():
        BANK_FACTORS_FILE.write_text(
            json.dumps(DEFAULT_BANK_FACTORS, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    if not BANK_LOGOS_FILE.exists():
        BANK_LOGOS_FILE.write_text("{}", encoding="utf-8")
    if not BANK_CATALOG_FILE.exists():
        default_banks = sorted({item["bank"] for item in DEFAULT_BANK_FACTORS}, key=str.lower)
        BANK_CATALOG_FILE.write_text(
            json.dumps(default_banks, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


def load_discounts() -> list[dict]:
    ensure_data_files()
    try:
        return json.loads(DISCOUNTS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return DEFAULT_DISCOUNTS


def load_discounts_raw() -> str:
    ensure_data_files()
    try:
        return DISCOUNTS_FILE.read_text(encoding="utf-8")
    except OSError:
        return json.dumps(DEFAULT_DISCOUNTS, ensure_ascii=False, indent=2)


def save_discounts(items: list[dict]) -> None:
    ensure_data_files()
    normalized = []
    for item in items:
        if not isinstance(item, dict):
            raise ValueError("Cada item precisa ser um objeto com 'name' e 'percent'.")
        name = str(item.get("name", "")).strip()
        if not name:
            continue
        percent = float(item.get("percent", 0) or 0)
        normalized.append({"name": name, "percent": percent})

    DISCOUNTS_FILE.write_text(
        json.dumps(normalized, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_bank_factors() -> list[dict]:
    ensure_data_files()
    try:
        items = json.loads(BANK_FACTORS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        items = DEFAULT_BANK_FACTORS
    return normalize_bank_factors(items)


def save_bank_factors(items: list[dict]) -> None:
    ensure_data_files()
    BANK_FACTORS_FILE.write_text(
        json.dumps(normalize_bank_factors(items), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_bank_catalog() -> list[str]:
    ensure_data_files()
    try:
        raw = json.loads(BANK_CATALOG_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        raw = sorted({item["bank"] for item in DEFAULT_BANK_FACTORS}, key=str.lower)
    if not isinstance(raw, list):
        return []
    names = sorted(
        {str(item).strip() for item in raw if str(item).strip()},
        key=str.lower,
    )
    return names


def save_bank_catalog(items: list[str]) -> None:
    ensure_data_files()
    names = sorted({str(item).strip() for item in items if str(item).strip()}, key=str.lower)
    BANK_CATALOG_FILE.write_text(
        json.dumps(names, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def register_bank(bank: str) -> None:
    cleaned = str(bank or "").strip()
    if not cleaned:
        raise ValueError("Informe o nome do banco.")
    banks = load_bank_catalog()
    if cleaned not in banks:
        banks.append(cleaned)
        save_bank_catalog(banks)


def load_blocked_entities() -> list[str]:
    try:
        raw = json.loads(BLOCKED_ENTITIES_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError, FileNotFoundError):
        return []
    if not isinstance(raw, list):
        return []
    return sorted(
        {str(item).strip() for item in raw if str(item).strip()},
        key=str.lower,
    )


def save_blocked_entities(items: list[str]) -> None:
    cleaned = sorted(
        {str(item).strip() for item in items if str(item).strip()},
        key=str.lower,
    )
    BLOCKED_ENTITIES_FILE.write_text(
        json.dumps(cleaned, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_bank_logos() -> dict[str, str]:
    ensure_data_files()
    try:
        raw = json.loads(BANK_LOGOS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    if not isinstance(raw, dict):
        return {}
    return {str(key): str(value) for key, value in raw.items() if key and value}


def save_bank_logos(items: dict[str, str]) -> None:
    ensure_data_files()
    BANK_LOGOS_FILE.write_text(
        json.dumps(items, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def get_bank_logo_url(bank: str, logos: dict[str, str] | None = None) -> str:
    if logos is None:
        logos = load_bank_logos()
    filename = logos.get(bank, "").strip()
    if not filename:
        return ""
    return url_for("static", filename=f"bank-logos/{filename}")


def save_uploaded_bank_logo(bank: str, file_storage) -> None:
    if not bank.strip() or not file_storage or not file_storage.filename:
        return

    ensure_data_files()
    extension = Path(file_storage.filename).suffix.lower()
    if extension not in ALLOWED_LOGO_EXTENSIONS:
        raise ValueError("Formato de logo invalido. Use PNG, JPG, JPEG ou WEBP.")

    slug = slugify_bank_name(bank)
    filename = secure_filename(f"{slug}{extension}")
    target = BANK_LOGOS_DIR / filename

    logos = load_bank_logos()
    previous = logos.get(bank)
    if previous and previous != filename:
        previous_path = BANK_LOGOS_DIR / previous
        if previous_path.exists():
            previous_path.unlink()

    file_storage.save(target)
    logos[bank] = filename
    save_bank_logos(logos)


def build_editor_bank_groups(product: str) -> list[dict]:
    logos = load_bank_logos()
    grouped: dict[str, list[dict]] = {}
    for item in load_bank_factors():
        if item.get("product", "normal") != product:
            continue
        grouped.setdefault(item["bank"], []).append(item)

    groups = []
    all_banks = sorted(set(load_bank_catalog()) | set(grouped.keys()) | set(logos.keys()), key=str.lower)
    for bank in all_banks:
        factors = sorted(grouped.get(bank, []), key=lambda item: -item["installments"])
        groups.append(
            {
                "bank": bank,
                "logo_url": get_bank_logo_url(bank, logos),
                "factors": factors,
                "factor_count": len(factors),
                "product": product,
            }
        )
    return groups


@app.route("/")
def index():
    return render_template("index.html", current_user=g.current_user)


@app.get("/admin")
@require_role("admin")
def admin_dashboard():
    react_build_dir = BASE_DIR / "static" / "react"
    if not (react_build_dir / "admin.html").exists():
        return "A interface administrativa ainda não foi compilada.", 503
    return send_from_directory(react_build_dir, "admin.html")


@app.get("/login")
def login():
    if session.get("user_id"):
        user = auth_store.get_user(session["user_id"])
        if user:
            return redirect(url_for("admin_dashboard") if user["role"] == "admin" else url_for("index"))
    react_build_dir = BASE_DIR / "static" / "react"
    if not (react_build_dir / "index.html").exists():
        return "A interface de login ainda não foi compilada. Execute: npm --prefix client run build", 503
    return send_from_directory(react_build_dir, "index.html")


@app.post("/api/auth/login")
def api_login():
    ip = request_ip()
    allowed, retry_after = presence_store.check_rate_limit(ip)
    if not allowed:
        auth_store.log("login_rate_limited", ip, details=f"Bloqueado por {retry_after}s")
        return jsonify({"error": f"Muitas tentativas. Tente novamente em {retry_after} segundos."}), 429

    payload = request.get_json(silent=True) or {}
    email = str(payload.get("email", ""))
    password = str(payload.get("password", ""))
    user = auth_store.verify_credentials(email, password)
    if not user:
        presence_store.record_login_attempt(ip)
        auth_store.log("login_failed", ip, details="Credenciais inválidas")
        return api_auth_error("E-mail ou senha inválidos.", 401)

    presence_store.clear_login_attempts(ip)

    if user.get("must_change_password"):
        session.clear()
        session["pending_password_change_user_id"] = user["id"]
        auth_store.log("login_must_change_password", ip, user["id"])
        return jsonify({"must_change_password": True})

    session.clear()
    session["user_id"] = user["id"]
    session.permanent = bool(payload.get("keep_connected"))
    auth_store.touch_user(user["id"])
    presence_store.mark_online(user["id"])
    auth_store.log("login_succeeded", ip, user["id"])
    redirect_to = url_for("admin_dashboard") if user["role"] == "admin" else url_for("index")
    return jsonify({"user": {"email": user["email"], "role": user["role"]}, "redirect_to": redirect_to})


@app.get("/api/auth/session")
def api_session():
    user_id = session.get("user_id")
    user = auth_store.get_user(user_id) if user_id else None
    if not user:
        return api_auth_error("Autenticação necessária.", 401)
    return jsonify({
        "user": {"id": user["id"], "email": user["email"], "role": user["role"]},
        "csrf_token": generate_csrf_token(),
    })


@app.post("/api/auth/change-password")
def api_change_password():
    user_id = session.get("pending_password_change_user_id")
    if not user_id:
        return api_auth_error("Sessão inválida. Faça login novamente.", 401)
    payload = request.get_json(silent=True) or {}
    new_password = str(payload.get("new_password", ""))
    confirmation = str(payload.get("new_password_confirmation", ""))
    if new_password != confirmation:
        return api_auth_error("A senha e a confirmação não conferem.", 400)
    try:
        auth_store.update_user(user_id, password=new_password, must_change_password=False)
    except ValueError as error:
        return api_auth_error(str(error), 400)
    session.clear()
    session["user_id"] = user_id
    auth_store.touch_user(user_id)
    presence_store.mark_online(user_id)
    auth_store.log("password_changed_first_login", request_ip(), user_id)
    user = auth_store.get_user(user_id)
    redirect_to = url_for("admin_dashboard") if user and user["role"] == "admin" else url_for("index")
    return jsonify({"redirect_to": redirect_to})


@app.post("/api/auth/logout")
def api_logout():
    user_id = session.get("user_id")
    if user_id:
        auth_store.log("logout", request_ip(), user_id)
        auth_store.mark_user_offline(user_id)
        presence_store.mark_offline(user_id)
    session.clear()
    return "", 204


@app.get("/api/admin/audit")
@require_role("admin")
def api_admin_audit():
    return jsonify(
        auth_store.list_audit_logs(
            page=request.args.get("page", 1, type=int),
            per_page=request.args.get("per_page", 25, type=int),
            event=request.args.get("event", "").strip(),
            search=request.args.get("search", "").strip(),
        )
    )


@app.get("/api/admin/summary")
@require_role("admin")
def api_admin_summary():
    return jsonify(auth_store.audit_summary())


@app.get("/api/admin/activity")
@require_role("admin")
def api_admin_activity():
    users = auth_store.list_users()
    online_user_ids = presence_store.online_user_ids([user["id"] for user in users])
    return jsonify(auth_store.user_activity(online_user_ids))


@app.get("/api/admin/activity/stream")
@require_role("admin")
def api_admin_activity_stream():
    @stream_with_context
    def event_stream():
        yield "data: {\"event\":\"connected\"}\n\n"
        for message in presence_store.listen():
            if message is None:
                yield ": keep-alive\n\n"
            else:
                yield f"data: {message}\n\n"

    return Response(
        event_stream(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/api/auth/heartbeat")
def api_auth_heartbeat():
    user_id = session.get("user_id")
    user = auth_store.get_user(user_id) if user_id else None
    if not user:
        return api_auth_error("Autenticação necessária.", 401)
    auth_store.touch_user(user["id"])
    presence_store.mark_online(user["id"])
    return "", 204


@app.get("/api/admin/users")
@require_role("admin")
def api_admin_users():
    return jsonify({"users": auth_store.list_users()})


@app.post("/api/admin/users")
@require_role("admin")
def api_admin_create_user():
    payload = request.get_json(silent=True) or {}
    email = str(payload.get("email", ""))
    role = str(payload.get("role", "operador"))
    display_name = str(payload.get("display_name", ""))
    password = str(payload.get("password", ""))
    if password != str(payload.get("password_confirmation", "")):
        return api_auth_error("A senha e a confirmação não conferem.", 400)
    try:
        user_id = auth_store.create_user(email, password, role, display_name, must_change_password=True)
    except ValueError as error:
        return api_auth_error(str(error), 400)

    auth_store.log(
        "user_created",
        request_ip(),
        g.current_user["id"],
        details=f"Usuário {email.strip().lower()} criado como {role}",
    )
    presence_store.publish("users_changed", user_id)
    return jsonify({"user": auth_store.get_user(user_id)}), 201


@app.post("/api/admin/users/<int:user_id>/reset-password")
@require_role("admin")
def api_admin_reset_password(user_id: int):
    temp_password = secrets.token_urlsafe(16)
    try:
        user = auth_store.update_user(user_id, password=temp_password, must_change_password=True)
    except ValueError as error:
        return api_auth_error(str(error), 400)

    auth_store.log(
        "password_reset",
        request_ip(),
        g.current_user["id"],
        details=f"Senha redefinida para {user['email']}",
    )
    return jsonify({"temporary_password": temp_password, "user": user})


@app.patch("/api/admin/users/<int:user_id>")
@require_role("admin")
def api_admin_update_user(user_id: int):
    payload = request.get_json(silent=True) or {}
    if user_id == g.current_user["id"] and payload.get("is_active") is False:
        return api_auth_error("Você não pode desativar sua própria conta.", 400)
    if user_id == g.current_user["id"] and payload.get("role") not in (None, "admin"):
        return api_auth_error("Você não pode remover seu próprio perfil administrativo.", 400)

    changes = {}
    if "email" in payload:
        changes["email"] = str(payload["email"])
    if "role" in payload:
        changes["role"] = str(payload["role"])
    if "is_active" in payload:
        changes["is_active"] = bool(payload["is_active"])
    if "display_name" in payload:
        changes["display_name"] = str(payload["display_name"])
    if payload.get("password"):
        if str(payload["password"]) != str(payload.get("password_confirmation", "")):
            return api_auth_error("A senha e a confirmação não conferem.", 400)
        changes["password"] = str(payload["password"])
    try:
        user = auth_store.update_user(user_id, **changes)
    except ValueError as error:
        return api_auth_error(str(error), 400)

    auth_store.log(
        "user_updated",
        request_ip(),
        g.current_user["id"],
        details=f"Usuário {user['email']} atualizado",
    )
    presence_store.publish("users_changed", user_id)
    return jsonify({"user": user})


@app.post("/api/admin/users/bulk")
@require_role("admin")
def api_admin_bulk_create_users():
    payload = request.get_json(silent=True) or {}
    rows = payload.get("users", [])
    if not isinstance(rows, list) or not rows:
        return api_auth_error("Envie uma lista de usuários.", 400)
    if len(rows) > 200:
        return api_auth_error("Máximo de 200 usuários por importação.", 400)

    created = []
    errors = []
    for idx, row in enumerate(rows, start=1):
        email = str(row.get("email", "")).strip()
        display_name = str(row.get("display_name", "")).strip()
        role = str(row.get("role", "operador")).strip().lower()
        password = str(row.get("password", "")).strip()
        if not password:
            password = secrets.token_urlsafe(16)
        try:
            user_id = auth_store.create_user(
                email, password, role, display_name, must_change_password=True,
            )
            created.append({"line": idx, "email": email, "display_name": display_name, "password": password})
        except (ValueError, Exception) as error:
            errors.append({"line": idx, "email": email, "error": str(error)})

    if created:
        auth_store.log(
            "users_bulk_created",
            request_ip(),
            g.current_user["id"],
            details=f"{len(created)} usuário(s) criado(s) em lote",
        )
        presence_store.publish("users_changed", 0)
    return jsonify({"created": created, "errors": errors})


@app.get("/api/discounts")
def api_discounts():
    return jsonify(load_discounts())


@app.get("/api/bank-factors")
def api_bank_factors():
    logos = load_bank_logos()
    return jsonify(
        [
            {
                **item,
                "logo_url": get_bank_logo_url(item["bank"], logos),
            }
            for item in load_bank_factors()
        ]
    )


@app.get("/api/commercial-matrix")
def api_commercial_matrix():
    try:
        rows = commercial_matrix_service.list_rows(
            bank=request.args.get("banco"),
            operation=request.args.get("operacao"),
        )
    except SheetsConfigurationError as error:
        return api_auth_error(str(error), 503)
    except HttpError:
        return api_auth_error("Não foi possível consultar a Matriz_Comercial.", 502)

    auth_store.log("commercial_matrix_read", request_ip(), g.current_user["id"])
    return jsonify({"total": len(rows), "rows": rows})


@app.get("/api/commercial-matrix/filters")
def api_commercial_matrix_filters():
    try:
        return jsonify(commercial_matrix_service.list_filters())
    except SheetsConfigurationError as error:
        return api_auth_error(str(error), 503)
    except HttpError:
        return api_auth_error("Não foi possível consultar a Matriz_Comercial.", 502)


@app.get("/api/blocked-entities")
def api_blocked_entities_get():
    return jsonify({"blocked": load_blocked_entities()})


@app.put("/api/admin/blocked-entities")
@require_role("admin")
def api_blocked_entities_put():
    payload = request.get_json(silent=True) or {}
    items = payload.get("blocked", [])
    if not isinstance(items, list):
        return api_auth_error("Formato inválido.", 400)
    save_blocked_entities(items)
    auth_store.log(
        "blocked_entities_updated",
        request_ip(),
        g.current_user["id"],
        details=f"{len(load_blocked_entities())} entidade(s) bloqueada(s)",
    )
    return jsonify({"blocked": load_blocked_entities()})


@app.get("/api/admin/reports/productivity")
@require_role("admin")
def api_admin_productivity_report():
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    if not date_from or not date_to:
        return api_auth_error("Informe os parâmetros 'from' e 'to'.", 400)
    rows = auth_store.productivity_report(date_from, date_to)
    return jsonify({"rows": rows})


@app.get("/api/admin/reports/productivity/excel")
@require_role("admin")
def api_admin_productivity_excel():
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    if not date_from or not date_to:
        return api_auth_error("Informe os parâmetros 'from' e 'to'.", 400)
    rows = auth_store.productivity_report(date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    headers = ["Login", "Consultor", "Data", "Primeiro relatório", "Último relatório", "TMs", "Simulações"]
    header_fill = PatternFill(start_color="FF9D1A", end_color="FF9D1A", fill_type="solid")
    header_font = Font(bold=True, color="161616", size=11)
    for col_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, 2):
        avg_seconds = row.get("avg_seconds")
        if avg_seconds is not None:
            minutes = int(avg_seconds) // 60
            seconds = int(avg_seconds) % 60
            tms = f"{minutes}m{seconds:02d}s"
        else:
            tms = "-"
        first_at = row.get("first_at")
        last_at = row.get("last_at")
        ws.cell(row=row_index, column=1, value=row.get("email", ""))
        ws.cell(row=row_index, column=2, value=row.get("display_name", ""))
        ws.cell(row=row_index, column=3, value=str(row.get("report_date", "")))
        ws.cell(row=row_index, column=4, value=first_at.strftime("%H:%M:%S") if first_at else "")
        ws.cell(row=row_index, column=5, value=last_at.strftime("%H:%M:%S") if last_at else "")
        ws.cell(row=row_index, column=6, value=tms)
        ws.cell(row=row_index, column=7, value=row.get("simulations", 0))

    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 35)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    auth_store.log("report_exported", request_ip(), g.current_user["id"], details=f"Relatório {date_from} a {date_to}")
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{date_from}_{date_to}.xlsx"},
    )


@app.post("/api/parse")
def api_parse():
    payload = request.get_json(silent=True) or {}
    rows = parse_table_text(payload.get("text", ""))
    auth_store.log(
        "table_processed",
        request_ip(),
        g.current_user["id"],
        details=f"{len(rows)} contrato(s) processado(s)",
    )
    presence_store.publish("table_processed", g.current_user["id"])
    return jsonify({"rows": rows})


@app.post("/api/export/excel")
def api_export_excel():
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows", [])
    if not rows:
        return api_auth_error("Nenhum dado para exportar.", 400)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    headers = [
        "Consignatária", "Situação", "ADE", "Serviço", "Prestações",
        "Pagas", "Prestação (R$)", "Deferimento", "Restantes",
        "Saldo Devedor (R$)", "Maturidade", "Banco — Operação",
        "Referência", "Valor previsão (R$)",
    ]
    header_fill = PatternFill(start_color="FF9D1A", end_color="FF9D1A", fill_type="solid")
    header_font = Font(bold=True, color="161616", size=11)
    for col_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, 2):
        ws.cell(row=row_index, column=1, value=row.get("consignataria", ""))
        ws.cell(row=row_index, column=2, value=row.get("situacao", ""))
        ws.cell(row=row_index, column=3, value=row.get("ade", ""))
        ws.cell(row=row_index, column=4, value=row.get("servico", ""))
        ws.cell(row=row_index, column=5, value=row.get("prestacoes"))
        ws.cell(row=row_index, column=6, value=row.get("pagas"))
        ws.cell(row=row_index, column=7, value=row.get("prestacao"))
        ws.cell(row=row_index, column=8, value=row.get("deferimento", ""))
        ws.cell(row=row_index, column=9, value=row.get("remaining"))
        ws.cell(row=row_index, column=10, value=row.get("debtBalance"))
        ws.cell(row=row_index, column=11, value=row.get("maturity", ""))
        ws.cell(row=row_index, column=12, value=row.get("commercialLabel", ""))
        ws.cell(row=row_index, column=13, value=row.get("reference", ""))
        ws.cell(row=row_index, column=14, value=row.get("forecastValue"))

    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 35)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    auth_store.log("export_excel", request_ip(), g.current_user["id"], details=f"{len(rows)} registros exportados")
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="calculadora-resultado.xlsx",
    )


@app.route("/discounts-editor", methods=["GET", "POST"])
@require_role("admin")
def discounts_editor():
    error = None
    if request.method == "POST":
        try:
            form_type = request.form.get("form_type", "discounts")
            if form_type == "add_bank":
                bank_name = request.form.get("bank_name", "").strip()
                register_bank(bank_name)
                save_uploaded_bank_logo(bank_name, request.files.get("bank_logo"))
            elif form_type == "modal_bank_logo":
                bank_name = request.form.get("modal_bank_name", "").strip()
                register_bank(bank_name)
                save_uploaded_bank_logo(bank_name, request.files.get("modal_bank_logo"))
            elif form_type == "bank_logos":
                for bank in request.form.getlist("logo_bank"):
                    save_uploaded_bank_logo(bank, request.files.get(f"logo_file_{bank}"))
            elif form_type == "bank_factors_payload":
                payload = json.loads(request.form.get("bank_factors_payload", "[]") or "[]")
                if not isinstance(payload, list):
                    raise ValueError("Payload de fatores invalido.")
                items = []
                banks = []
                for group in payload:
                    if not isinstance(group, dict):
                        continue
                    bank = str(group.get("bank", "")).strip()
                    if not bank:
                        continue
                    product = str(group.get("product", "normal") or "normal").strip().lower()
                    if product not in {"normal", "tj"}:
                        product = "normal"
                    banks.append(bank)
                    factors = group.get("factors", [])
                    if not isinstance(factors, list):
                        continue
                    for factor_item in factors:
                        if not isinstance(factor_item, dict):
                            continue
                        items.append(
                            {
                                "bank": bank,
                                "installments": factor_item.get("installments", 0),
                                "factor": factor_item.get("factor", 0),
                                "active": True,
                                "product": product,
                            }
                        )
                save_bank_catalog(load_bank_catalog() + banks)
                save_bank_factors(items)
                auth_store.log("bank_factors_updated", request_ip(), g.current_user["id"], details=f"{len(items)} fator(es) atualizado(s)")
            elif form_type == "bank_factors":
                banks = request.form.getlist("factor_bank")
                installments = request.form.getlist("factor_installments")
                factors = request.form.getlist("factor_value")
                items = []
                for index, (bank, installment, factor) in enumerate(
                    zip(banks, installments, factors)
                ):
                    cleaned_bank = bank.strip()
                    if not cleaned_bank:
                        continue
                    items.append(
                        {
                            "bank": cleaned_bank,
                            "installments": int(float(str(installment or "0").replace(",", "."))),
                            "factor": float(str(factor or "0").replace(",", ".")),
                            "active": True,
                        }
                    )
                save_bank_factors(items)
                auth_store.log("bank_factors_updated", request_ip(), g.current_user["id"], details=f"{len(items)} fator(es) atualizado(s)")
            else:
                names = request.form.getlist("discount_name")
                percents = request.form.getlist("discount_percent")
                items = []
                for name, percent in zip(names, percents):
                    cleaned_name = name.strip()
                    if not cleaned_name:
                        continue
                    items.append(
                        {
                            "name": cleaned_name,
                            "percent": float(str(percent or "0").replace(",", ".")),
                        }
                    )
                save_discounts(items)
                auth_store.log("discounts_updated", request_ip(), g.current_user["id"], details=f"{len(items)} desconto(s) atualizado(s)")
            return redirect(url_for("discounts_editor", saved="1"))
        except ValueError as exc:
            error = str(exc)

    return render_template(
        "discounts_editor.html",
        discounts=load_discounts(),
        bank_groups_normal=build_editor_bank_groups("normal"),
        bank_groups_tj=build_editor_bank_groups("tj"),
        saved=request.args.get("saved") == "1",
        error=error,
        current_user=g.current_user,
    )


if __name__ == "__main__":
    ensure_data_files()
    app.run(host="0.0.0.0", port=5000, debug=True)
